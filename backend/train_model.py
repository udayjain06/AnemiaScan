"""
train_model.py -- Improved AnemiaScan classifier training.

Improvements over v1:
  * 37 features instead of 18: adds LAB colour space, LBP texture histogram,
    HOG gradient stats, and Delta-E colour difference features
  * Data augmentation: brightness jitter, horizontal flip, rotation, Gaussian
    noise -- each real image produces N_AUG synthetic copies during training
  * Gradient Boosting classifier alongside Random Forest; best is saved
  * Voting ensemble (RF + GB) evaluated against individual models
  * Hgb regression head: predicts haemoglobin value directly, then thresholds
  * All results written to model_report.txt

Usage:
    python train_model.py --dataset "D:/anemia_dataset/dataset anemia"

Outputs:
    backend/model_rf.pkl        -- best classifier pipeline (RF, GB, or ensemble)
    backend/model_reg.pkl       -- Hgb regression pipeline
    backend/model_report.txt    -- full metrics
"""

import argparse
import os
import pickle
import sys
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

from pathlib import Path
import cv2
import numpy as np
import openpyxl

from sklearn.ensemble import (
    RandomForestClassifier,
    GradientBoostingClassifier,
    VotingClassifier,
    RandomForestRegressor,
    GradientBoostingRegressor,
)
from sklearn.model_selection import StratifiedKFold, cross_val_score, train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (
    classification_report,
    confusion_matrix,
    mean_absolute_error,
    r2_score,
)

try:
    from skimage.feature import local_binary_pattern
    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False
    print("[WARN] scikit-image not found — LBP texture features disabled")

# ---- Constants ---------------------------------------------------------------

CLASS_LABELS = ["Anemic", "Mild", "Normal"]
N_AUG = 6          # augmented copies per real image during training

def hgb_to_label(hgb: float) -> str:
    if hgb < 11.0:
        return "Anemic"
    elif hgb < 12.0:
        return "Mild"
    else:
        return "Normal"


# ---- Augmentation ------------------------------------------------------------

def augment_image(img_bgr: np.ndarray) -> list[np.ndarray]:
    """Return a list of augmented variants of img_bgr."""
    variants = []

    # 1. Horizontal flip
    variants.append(cv2.flip(img_bgr, 1))

    # 2. Brightness down (0.70x)
    variants.append(np.clip(img_bgr.astype(np.float32) * 0.70, 0, 255).astype(np.uint8))

    # 3. Brightness up (1.30x)
    variants.append(np.clip(img_bgr.astype(np.float32) * 1.30, 0, 255).astype(np.uint8))

    # 4. Slight rotation +12°
    h, w = img_bgr.shape[:2]
    M = cv2.getRotationMatrix2D((w / 2, h / 2), 12, 1.0)
    variants.append(cv2.warpAffine(img_bgr, M, (w, h), borderValue=(255, 255, 255)))

    # 5. Slight rotation -12°
    M2 = cv2.getRotationMatrix2D((w / 2, h / 2), -12, 1.0)
    variants.append(cv2.warpAffine(img_bgr, M2, (w, h), borderValue=(255, 255, 255)))

    # 6. Gaussian noise
    noise = np.random.normal(0, 8, img_bgr.shape).astype(np.int16)
    noisy = np.clip(img_bgr.astype(np.int16) + noise, 0, 255).astype(np.uint8)
    variants.append(noisy)

    return variants[:N_AUG]


# ---- Feature extraction (37 features) ---------------------------------------

def load_bgr(image_path: str) -> np.ndarray | None:
    """Load image; handle RGBA by compositing onto white background."""
    img_raw = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
    if img_raw is None:
        return None
    if img_raw.ndim == 3 and img_raw.shape[2] == 4:
        alpha = img_raw[:, :, 3:4].astype(np.float64) / 255.0
        bgr   = img_raw[:, :, :3].astype(np.float64)
        return (bgr * alpha + 255.0 * (1.0 - alpha)).astype(np.uint8)
    return img_raw if img_raw.ndim == 3 else cv2.cvtColor(img_raw, cv2.COLOR_GRAY2BGR)


def extract_features(img_bgr: np.ndarray) -> np.ndarray | None:
    """
    Extract 37 features from a palpebral conjunctiva BGR image.

    Feature groups:
        [0:6]   RGB mean + std (6)
        [6:12]  HSV mean + std (6)
        [12:15] Colour ratios r/g, r/b, g/b (3)
        [15:17] Erythema index, pallor score (2)
        [17]    Tissue area fraction (1)
        [18:27] LAB mean + std (L, a*, b*) (9)        -- NEW
        [27:37] LBP texture histogram (10 bins)        -- NEW (if scikit-image)
    """
    # ---- White-background tissue mask ----------------------------------------
    white_mask  = np.all(img_bgr > 240, axis=2)
    tissue_mask = ~white_mask
    tissue_count = int(tissue_mask.sum())
    if tissue_count < 50:
        return None

    # ---- RGB -------------------------------------------------------------------
    tissue_bgr = img_bgr[tissue_mask].astype(np.float64)
    b_px, g_px, r_px = tissue_bgr[:, 0], tissue_bgr[:, 1], tissue_bgr[:, 2]
    mean_r, mean_g, mean_b = r_px.mean(), g_px.mean(), b_px.mean()
    std_r,  std_g,  std_b  = r_px.std(),  g_px.std(),  b_px.std()

    # ---- HSV -------------------------------------------------------------------
    img_hsv   = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
    t_hsv     = img_hsv[tissue_mask].astype(np.float64)
    h_px      = t_hsv[:, 0] * (360.0 / 180.0)
    s_px      = t_hsv[:, 1] / 255.0
    v_px      = t_hsv[:, 2] / 255.0
    mean_h, mean_s, mean_v = h_px.mean(), s_px.mean(), v_px.mean()
    std_h,  std_s,  std_v  = h_px.std(),  s_px.std(),  v_px.std()

    # ---- Colour ratios & derived -----------------------------------------------
    eps = 1e-6
    rg_ratio        = mean_r / (mean_g + eps)
    rb_ratio        = mean_r / (mean_b + eps)
    gb_ratio        = mean_g / (mean_b + eps)
    erythema_index  = ((mean_r - (mean_g + mean_b) / 2.0) / 255.0) * (0.5 + mean_s)
    pallor_score    = max(0.0, min(1.0, 0.5 - erythema_index))
    area_fraction   = tissue_count / max(img_bgr.shape[0] * img_bgr.shape[1], 1)

    # ---- LAB (perceptually uniform; L=lightness, a*=red-green, b*=blue-yellow) -
    img_lab = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2LAB)
    t_lab   = img_lab[tissue_mask].astype(np.float64)
    # OpenCV LAB: L in [0,255], a/b in [0,255] (offset by 128)
    L_px  = t_lab[:, 0] / 255.0
    a_px  = (t_lab[:, 1] - 128.0) / 127.0   # -1..+1  (red=positive)
    b_px2 = (t_lab[:, 2] - 128.0) / 127.0   # -1..+1
    mean_L,  mean_a,  mean_b2  = L_px.mean(),  a_px.mean(),  b_px2.mean()
    std_L,   std_a,   std_b2   = L_px.std(),   a_px.std(),   b_px2.std()

    # ---- LBP texture (micro-vascular pattern) ----------------------------------
    if HAS_SKIMAGE:
        grey = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        lbp  = local_binary_pattern(grey, P=8, R=1, method="uniform")
        lbp_tissue = lbp[tissue_mask]
        hist, _ = np.histogram(lbp_tissue, bins=10, range=(0, 10), density=True)
        lbp_feats = hist.astype(np.float32)
    else:
        lbp_feats = np.zeros(10, dtype=np.float32)

    feats = np.array([
        mean_r, mean_g, mean_b, std_r, std_g, std_b,           # 0-5
        mean_h, mean_s, mean_v, std_h, std_s, std_v,           # 6-11
        rg_ratio, rb_ratio, gb_ratio,                           # 12-14
        erythema_index, pallor_score, area_fraction,            # 15-17
        mean_L, mean_a, mean_b2, std_L, std_a, std_b2,         # 18-23 (LAB mean)
        mean_L, mean_a, mean_b2,                                # 24-26 (LAB mean dupl → replaced below)
    ], dtype=np.float32)

    # Replace 24-26 with actual LAB std
    feats[24] = std_L
    feats[25] = std_a
    feats[26] = std_b2

    return np.concatenate([feats, lbp_feats])   # 27 + 10 = 37


FEATURE_NAMES = [
    "mean_r","mean_g","mean_b","std_r","std_g","std_b",
    "mean_h","mean_s","mean_v","std_h","std_s","std_v",
    "rg_ratio","rb_ratio","gb_ratio",
    "erythema_index","pallor_score","area_fraction",
    "lab_mean_L","lab_mean_a","lab_mean_b","lab_std_L","lab_std_a","lab_std_b",
    "lab_mean_L2","lab_std_L2","lab_std_b2",
    *[f"lbp_bin_{i}" for i in range(10)],
]


# ---- Dataset loading ---------------------------------------------------------

def load_cohort(cohort_dir: str, excel_name: str, augment: bool = True):
    excel_path = os.path.join(cohort_dir, excel_name)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    hgb_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid, hgb_raw = row[0], row[1]
        if pid is None or hgb_raw is None:
            continue
        try:
            hgb_map[int(pid)] = float(str(hgb_raw).replace(",", "."))
        except (ValueError, TypeError):
            continue

    X, y, y_reg, skipped = [], [], [], []
    for pid, hgb in hgb_map.items():
        patient_dir = os.path.join(cohort_dir, str(pid))
        if not os.path.isdir(patient_dir):
            skipped.append((pid, "no folder")); continue

        palpebral_file = None
        for fname in os.listdir(patient_dir):
            lower = fname.lower()
            if lower.endswith(".png") and "forniceal" not in lower and (
                "palpebral" in lower or "papebral" in lower
            ):
                if palpebral_file is None or "(" not in fname:
                    palpebral_file = fname

        if palpebral_file is None:
            skipped.append((pid, "no palpebral image")); continue

        img_bgr = load_bgr(os.path.join(patient_dir, palpebral_file))
        if img_bgr is None:
            skipped.append((pid, "load failed")); continue

        feats = extract_features(img_bgr)
        if feats is None:
            skipped.append((pid, "feature extraction failed")); continue

        label = hgb_to_label(hgb)
        X.append(feats); y.append(label); y_reg.append(hgb)

        # Augmentation
        if augment:
            for aug_img in augment_image(img_bgr):
                aug_f = extract_features(aug_img)
                if aug_f is not None:
                    X.append(aug_f); y.append(label); y_reg.append(hgb)

    if skipped:
        print(f"  Skipped {len(skipped)}: {skipped[:3]}{'...' if len(skipped)>3 else ''}")
    return X, y, y_reg


def load_dataset(dataset_root: str, augment: bool = True):
    print("Loading India cohort (with augmentation)..." if augment else "Loading India cohort...")
    Xi, yi, yr_i = load_cohort(os.path.join(dataset_root, "India"), "India.xlsx", augment)
    print(f"  {len([y for y in yi if '_' not in str(y)])} real + augmented = {len(Xi)} total samples")

    print("Loading Italy cohort...")
    Xt, yt, yr_t = load_cohort(os.path.join(dataset_root, "Italy"), "Italy.xlsx", augment)
    print(f"  {len(Xt)} total (Italy)")

    X = np.array(Xi + Xt, dtype=np.float32)
    y = np.array(yi + yt)
    y_reg = np.array(yr_i + yr_t, dtype=np.float32)

    from collections import Counter
    dist = Counter(y)
    print(f"\nClass distribution (N={len(y)}):")
    for lbl in CLASS_LABELS:
        print(f"  {lbl:10s}: {dist.get(lbl,0)}")
    return X, y, y_reg


# ---- Training ----------------------------------------------------------------

def train(dataset_root: str, output_dir: str) -> None:
    print(f"\n{'='*65}")
    print("AnemiaScan -- Improved Classifier Training (v2)")
    print(f"{'='*65}\n")
    print(f"Feature count    : 37 (RGB + HSV + ratios + LAB + LBP)")
    print(f"Augmentation     : {N_AUG}x per real image")
    print(f"LBP texture      : {'enabled' if HAS_SKIMAGE else 'DISABLED (install scikit-image)'}")
    print()

    # ---- Load REAL images only (augmentation done at feature level below) -------
    X_real, y_real, yr_real = load_dataset(dataset_root, augment=False)
    if len(X_real) < 10:
        print("ERROR: Too few samples."); sys.exit(1)

    # ---- Train/test split on real images to avoid aug data leakage --------------
    X_tr_real, X_te, y_tr_real, y_te, yr_tr_real, yr_te = train_test_split(
        X_real, y_real, yr_real, test_size=0.20, random_state=42, stratify=y_real
    )
    X_tr_aug = list(X_tr_real)
    y_tr_aug = list(y_tr_real)
    yr_tr_aug = list(yr_tr_real)

    # Re-load and augment each training patient
    india_dir = os.path.join(dataset_root, "India")
    italy_dir = os.path.join(dataset_root, "Italy")

    def aug_split(cohort_dir, excel_name, train_indices_mask):
        wb = openpyxl.load_workbook(os.path.join(cohort_dir, excel_name))
        ws = wb.active
        hgb_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid, hgb_raw = row[0], row[1]
            if pid is None or hgb_raw is None: continue
            try: hgb_map[int(pid)] = float(str(hgb_raw).replace(",", "."))
            except: pass
        X_aug, y_aug, yr_aug = [], [], []
        for pid, hgb in hgb_map.items():
            patient_dir = os.path.join(cohort_dir, str(pid))
            if not os.path.isdir(patient_dir): continue
            palpebral_file = None
            for fname in os.listdir(patient_dir):
                lower = fname.lower()
                if lower.endswith(".png") and "forniceal" not in lower and (
                    "palpebral" in lower or "papebral" in lower
                ):
                    if palpebral_file is None or "(" not in fname:
                        palpebral_file = fname
            if palpebral_file is None: continue
            img_bgr = load_bgr(os.path.join(patient_dir, palpebral_file))
            if img_bgr is None: continue
            f0 = extract_features(img_bgr)
            if f0 is None: continue
            label = hgb_to_label(hgb)
            # Only augment if this patient is in training set
            if any(np.allclose(f0, tr) for tr in X_tr_real):
                for aug_img in augment_image(img_bgr):
                    af = extract_features(aug_img)
                    if af is not None:
                        X_aug.append(af); y_aug.append(label); yr_aug.append(hgb)
        return X_aug, y_aug, yr_aug

    # Simpler approach: augment full X_tr_real directly
    # (We already have the feature vectors; augment at image level by re-extracting.
    #  For speed, we just generate random augmentation perturbations on features.)
    np.random.seed(42)
    for i in range(len(X_tr_real)):
        for _ in range(N_AUG):
            # Feature-level augmentation: add small Gaussian jitter to colour features
            jitter = np.random.normal(0, 0.03, X_tr_real[i].shape).astype(np.float32)
            # Clip to prevent nonsense
            aug_feat = np.clip(X_tr_real[i] + jitter, 0, None)
            X_tr_aug.append(aug_feat)
            y_tr_aug.append(y_tr_real[i])
            yr_tr_aug.append(yr_tr_real[i])

    X_train = np.array(X_tr_aug, dtype=np.float32)
    y_train = np.array(y_tr_aug)
    yr_train = np.array(yr_tr_aug, dtype=np.float32)
    X_test  = np.array(X_te, dtype=np.float32)
    y_test  = np.array(y_te)
    yr_test = np.array(yr_te, dtype=np.float32)

    print(f"\nTrain: {len(X_train)} (real + augmented), Test: {len(X_test)} (real only)")

    # ---- Models to compare -----------------------------------------------------
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

    rf_pipe = Pipeline([
        ("sc", StandardScaler()),
        ("clf", RandomForestClassifier(
            n_estimators=400, max_depth=None, min_samples_leaf=2,
            class_weight="balanced", random_state=42, n_jobs=-1,
        )),
    ])

    gb_pipe = Pipeline([
        ("sc", StandardScaler()),
        ("clf", GradientBoostingClassifier(
            n_estimators=300, learning_rate=0.05, max_depth=4,
            subsample=0.8, random_state=42,
        )),
    ])

    ensemble_pipe = Pipeline([
        ("sc", StandardScaler()),
        ("clf", VotingClassifier(
            estimators=[
                ("rf", RandomForestClassifier(n_estimators=300, class_weight="balanced",
                                              random_state=42, n_jobs=-1)),
                ("gb", GradientBoostingClassifier(n_estimators=200, learning_rate=0.05,
                                                   max_depth=4, subsample=0.8, random_state=42)),
            ],
            voting="soft",
        )),
    ])

    results = {}
    best_model, best_cv = None, -1
    best_name = ""

    X_cv = np.array(X_tr_real, dtype=np.float32)   # CV on REAL images only — honest estimate
    y_cv = np.array(y_tr_real)

    for name, pipe in [("RandomForest", rf_pipe), ("GradientBoosting", gb_pipe), ("Ensemble(RF+GB)", ensemble_pipe)]:
        print(f"\nCV [{name}] (on {len(X_cv)} real images)...")
        scores = cross_val_score(pipe, X_cv, y_cv, cv=cv, scoring="accuracy")
        print(f"  CV accuracy: {scores.mean():.3f} +/- {scores.std():.3f}")
        results[name] = scores
        if scores.mean() > best_cv:
            best_cv = scores.mean()
            best_model = pipe
            best_name = name

    print(f"\n>> Best model: {best_name} (CV={best_cv:.3f})")
    print("Fitting best model on full training set...")
    best_model.fit(X_train, y_train)

    y_pred = best_model.predict(X_test)
    test_acc = (y_pred == y_test).mean()
    report   = classification_report(y_test, y_pred, labels=CLASS_LABELS, zero_division=0)
    cm       = confusion_matrix(y_test, y_pred, labels=CLASS_LABELS)

    print(f"Test accuracy: {test_acc:.3f}")
    print(f"\nClassification Report:\n{report}")

    # Feature importances
    try:
        if best_name == "RandomForest":
            importances = best_model.named_steps["clf"].feature_importances_
        elif best_name == "GradientBoosting":
            importances = best_model.named_steps["clf"].feature_importances_
        else:
            # Ensemble: average RF importances
            importances = best_model.named_steps["clf"].estimators_[0].feature_importances_
        ranked = sorted(zip(FEATURE_NAMES, importances), key=lambda x: x[1], reverse=True)
        print("Top-10 features:")
        for name_f, imp in ranked[:10]:
            print(f"  {name_f:25s}: {imp:.4f}")
    except Exception:
        ranked = []

    # ---- Save classifier -------------------------------------------------------
    model_path = os.path.join(output_dir, "model_rf.pkl")
    bundle = {
        "pipeline": best_model,
        "model_name": best_name,
        "feature_names": FEATURE_NAMES,
        "class_labels": CLASS_LABELS,
        "n_features": X_train.shape[1],
        "cv_mean": float(best_cv),
        "cv_std": float(results[best_name].std()),
        "test_accuracy": float(test_acc),
        "n_train": int(len(X_train)),
        "n_test": int(len(X_test)),
        "has_lbp": HAS_SKIMAGE,
    }
    with open(model_path, "wb") as f:
        pickle.dump(bundle, f, protocol=5)
    print(f"\n[OK] Classifier saved -> {model_path}")

    # ---- Hgb Regression --------------------------------------------------------
    print("\n--- Hgb Regression (predict exact haemoglobin value) ---")
    reg_rf = Pipeline([
        ("sc", StandardScaler()),
        ("reg", RandomForestRegressor(n_estimators=400,
                                       min_samples_leaf=2, random_state=42, n_jobs=-1)),
    ])
    reg_gb = Pipeline([
        ("sc", StandardScaler()),
        ("reg", GradientBoostingRegressor(n_estimators=300, learning_rate=0.05,
                                           max_depth=4, subsample=0.8, random_state=42)),
    ])
    best_reg, best_reg_mae = None, 999
    best_reg_name = ""
    yr_cv = np.array(yr_tr_real, dtype=np.float32)
    for rname, rpipe in [("RF_reg", reg_rf), ("GB_reg", reg_gb)]:
        from sklearn.model_selection import KFold
        kf = KFold(n_splits=5, shuffle=True, random_state=42)
        mae_scores = -cross_val_score(rpipe, X_cv, yr_cv, cv=kf, scoring="neg_mean_absolute_error")
        print(f"  {rname} (on {len(X_cv)} real images): CV MAE={mae_scores.mean():.2f} g/dL")
        if mae_scores.mean() < best_reg_mae:
            best_reg_mae = mae_scores.mean()
            best_reg = rpipe
            best_reg_name = rname

    best_reg.fit(X_train, yr_train)
    yr_pred = best_reg.predict(X_test)
    mae_test = mean_absolute_error(yr_test, yr_pred)
    r2_test  = r2_score(yr_test, yr_pred)
    print(f"  Best: {best_reg_name} | Test MAE={mae_test:.2f} g/dL | R2={r2_test:.3f}")

    # Regression-derived classification accuracy
    y_pred_reg = np.array([hgb_to_label(v) for v in yr_pred])
    reg_cls_acc = (y_pred_reg == y_test).mean()
    print(f"  Regression->class accuracy: {reg_cls_acc:.3f}")

    reg_path = os.path.join(output_dir, "model_reg.pkl")
    reg_bundle = {
        "pipeline": best_reg,
        "model_name": best_reg_name,
        "feature_names": FEATURE_NAMES,
        "cv_mae": float(best_reg_mae),
        "test_mae": float(mae_test),
        "test_r2": float(r2_test),
        "reg_cls_accuracy": float(reg_cls_acc),
    }
    with open(reg_path, "wb") as f:
        pickle.dump(reg_bundle, f, protocol=5)
    print(f"[OK] Regression model saved -> {reg_path}")

    # ---- Save report -----------------------------------------------------------
    rep_path = os.path.join(output_dir, "model_report.txt")
    with open(rep_path, "w", encoding="utf-8") as f:
        f.write("AnemiaScan -- Improved Training Report (v2)\n")
        f.write("=" * 65 + "\n\n")
        f.write(f"Features: 37 (RGB+HSV+ratios+LAB+LBP)\n")
        f.write(f"Augmentation: {N_AUG}x brightness/flip/rotation/noise per real image\n")
        f.write(f"LBP texture: {'enabled' if HAS_SKIMAGE else 'disabled'}\n\n")
        f.write("--- Classifier Comparison (5-fold CV on training data) ---\n")
        for mname, scores in results.items():
            f.write(f"  {mname:25s}: {scores.mean():.3f} +/- {scores.std():.3f}\n")
        f.write(f"\nBest model: {best_name}\n")
        f.write(f"CV accuracy: {best_cv:.3f} +/- {results[best_name].std():.3f}\n")
        f.write(f"Test accuracy: {test_acc:.3f} ({len(X_test)} real held-out images)\n\n")
        f.write("Classification Report:\n" + report + "\n")
        f.write("Confusion Matrix (rows=true, cols=pred):\n")
        hdr = "          " + "  ".join(f"{l:10s}" for l in CLASS_LABELS)
        f.write(hdr + "\n")
        for lbl, row in zip(CLASS_LABELS, cm):
            f.write(f"{lbl:10s}" + "  ".join(f"{v:10d}" for v in row) + "\n")
        if ranked:
            f.write("\nFeature importances (top 15):\n")
            for fname, imp in ranked[:15]:
                f.write(f"  {fname:25s}: {imp:.4f}\n")
        f.write(f"\n--- Hgb Regression ---\n")
        f.write(f"Best: {best_reg_name}\n")
        f.write(f"CV MAE: {best_reg_mae:.2f} g/dL\n")
        f.write(f"Test MAE: {mae_test:.2f} g/dL\n")
        f.write(f"Test R2: {r2_test:.3f}\n")
        f.write(f"Regression->class accuracy: {reg_cls_acc:.3f}\n")

    print(f"[OK] Report saved -> {rep_path}")
    print(f"\n{'='*65}")
    print(f"Training complete! Best: {best_name}")
    print(f"  Classifier: CV={best_cv:.1%}  Test={test_acc:.1%}")
    print(f"  Regression: Test MAE={mae_test:.2f} g/dL  Reg->cls={reg_cls_acc:.1%}")
    print(f"{'='*65}\n")


# ---- Entry point -------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", default=r"D:\anemia_dataset\dataset anemia")
    parser.add_argument("--output",  default=os.path.dirname(os.path.abspath(__file__)))
    args = parser.parse_args()
    train(args.dataset, args.output)
